import openpyxl, os

F = lambda bold=False: openpyxl.styles.Font(bold=bold, color='FF000000')
dir_ = 'stick-world/config/excel'

# =============================================
# 扩张数据.xlsx: regions + factions + civ_diffs
# =============================================
wb = openpyxl.Workbook()

ws = wb.active; ws.title = "regions"
headers = ["id","name_zh","terrain","resources","stickmen_types","tech_unlocks","initial_owner","center_x","center_y"]
cn_rows = ["唯一ID","中文名","地形","资源(逗号分隔)","本地火柴人族群","可解锁科技","初始归属","中心X","中心Y"]
ws.append(headers); ws.append(cn_rows)
for c in range(1,10): ws.cell(row=1,column=c).font=F(True); ws.cell(row=2,column=c).font=F()
regions = [
    ["reg_plain_01","中原平原","plain","food,wood","plain","","","500","500"],
    ["reg_volcano_01","炎炉山脉","mountain","metal,black_pitch","volcano,giant","tech_smelting","","200","300"],
    ["reg_source_01","源流湿地","swamp","food,black_pitch","source","tech_magic_basic","","700","200"],
    ["reg_desert_01","赤砂荒原","desert","stone","desert,centaur","tech_camelry","","900","600"],
    ["reg_ocean_01","碧波群岛","coast","food,wood","ocean","tech_sailing","","300","800"],
    ["reg_forest_01","苍翠密林","forest","wood,food","forest,winged","tech_arbalest","","600","700"],
    ["reg_ice_01","极北冰原","ice","stone,black_pitch","ice","tech_frost_resist","","400","100"],
]
for row in regions: ws.append(row)
for c in range(1,10): ws.column_dimensions[chr(64+c)].width = 18

ws2 = wb.create_sheet("factions")
ws2.append(["id","name_zh","civilization_type","start_region","color_hex","description"])
ws2.append(["唯一ID","势力名","文明类型","起始地块","颜色","描述"])
for c in range(1,7): ws2.cell(row=1,column=c).font=F(True); ws2.cell(row=2,column=c).font=F()
factions = [
    ["fac_plain","平原王国","agriculture","reg_plain_01","#4CAF50","平原农耕文明，人口众多"],
    ["fac_volcano","炎炉部落","agriculture","reg_volcano_01","#FF5722","火山锻造文明，军工发达"],
    ["fac_source","源流学院","agriculture","reg_source_01","#9C27B0","法术文明，法师比例极高"],
    ["fac_desert","赤砂汗国","nomadic","reg_desert_01","#FFC107","游牧文明，骑兵精锐"],
    ["fac_ocean","碧波联邦","commerce","reg_ocean_01","#2196F3","海洋贸易文明，造船发达"],
    ["fac_forest","苍翠猎团","agriculture","reg_forest_01","#4CAF50","森林狩猎文明，隐匿强弓"],
    ["fac_ice","极北城邦","agriculture","reg_ice_01","#90CAF9","冰原生存专家，防御型文明"],
]
for row in factions: ws2.append(row)
for c in range(1,7): ws2.column_dimensions[chr(64+c)].width = 20

ws3 = wb.create_sheet("civ_differences")
ws3.append(["civilization_type","org_depth_bonus","military_bonus","economy_bonus","research_bonus","naval_bonus","description"])
ws3.append(["文明类型","组织层级倾向","军事加成","经济加成","科研加成","海军加成","描述"])
for c in range(1,8): ws3.cell(row=1,column=c).font=F(True); ws3.cell(row=2,column=c).font=F()
diffs = [
    ["agriculture","deep","balanced","中","中","低","农耕文明组织严密，中央集权"],
    ["nomadic","shallow","高(骑兵)","低","低","无","游牧来去如风，层级灵活"],
    ["commerce","medium","低","高","中","高","商业文明贸易驱动，军事靠外包"],
    ["ocean","medium","中","高(贸易)","中","极高","海洋文明制海权至上"],
]
for row in diffs: ws3.append(row)
for c in range(1,8): ws3.column_dimensions[chr(64+c)].width = 22

wb.save(f'{dir_}/扩张数据.xlsx'); wb.close(); print("扩张数据.xlsx done")

# =============================================
# 运输数据.xlsx: carriers + roads
# =============================================
wb2 = openpyxl.Workbook()

ws = wb2.active; ws.title = "carriers"
ws.append(["id","name_zh","speed","capacity","cost_per_km","terrain_penalty","unlock_tech"])
ws.append(["唯一ID","名称","速度","运量","单位成本","地形惩罚","解锁科技"])
for c in range(1,8): ws.cell(row=1,column=c).font=F(True); ws.cell(row=2,column=c).font=F()
carriers = [
    ["car_human","人力搬运","20","5","1","low","none"],
    ["car_beast","驮兽","40","15","3","medium","none"],
    ["car_wagon","马车","60","30","5","high","tech_wheel"],
    ["car_river","河船","80","50","10","water_only","tech_sailing"],
    ["car_ship","海船","100","200","20","ocean_only","tech_seafaring"],
    ["car_magic","魔法传送阵","999","10","50","none","tech_teleport"],
]
for row in carriers: ws.append(row)
for c in range(1,8): ws.column_dimensions[chr(64+c)].width = 18

ws2 = wb2.create_sheet("roads")
ws2.append(["id","name_zh","speed_bonus","build_cost_stone","build_cost_wood","build_time","maintenance","unlock_tech"])
ws2.append(["唯一ID","名称","速度加成","石料消耗","木材消耗","建造时间","维护成本","解锁科技"])
for c in range(1,9): ws2.cell(row=1,column=c).font=F(True); ws2.cell(row=2,column=c).font=F()
roads = [
    ["road_dirt","土路","1.2x","20","10","5","1","none"],
    ["road_stone","石板路","1.5x","80","20","15","3","tech_stone_road"],
    ["road_highway","帝国大道","2.0x","200","50","30","8","tech_imperial_road"],
    ["canal","运河","3.0x","500","0","60","15","tech_canal"],
]
for row in roads: ws2.append(row)
for c in range(1,9): ws2.column_dimensions[chr(64+c)].width = 18

wb2.save(f'{dir_}/运输数据.xlsx'); wb2.close(); print("运输数据.xlsx done")

# =============================================
# 编制预设: add engineering, admin, commerce
# =============================================
path = f'{dir_}/编制预设.xlsx'
wb3 = openpyxl.load_workbook(path)
ws3 = wb3.worksheets[0]

new_presets = [
    # Engineering
    ["tier_eng_hq","org_preset_engineering","工程总署",5,""],
    ["tier_eng_bureau","org_preset_engineering","工程司",4,"tier_eng_hq"],
    ["tier_eng_project","org_preset_engineering","项目组",3,"tier_eng_bureau"],
    ["tier_eng_team","org_preset_engineering","施工队",2,"tier_eng_project"],
    ["tier_eng_squad","org_preset_engineering","工匠班",1,"tier_eng_team"],
    # Administration
    ["tier_adm_court","org_preset_admin","朝廷",5,""],
    ["tier_adm_state","org_preset_admin","州",4,"tier_adm_court"],
    ["tier_adm_region","org_preset_admin","郡",3,"tier_adm_state"],
    ["tier_adm_county","org_preset_admin","县",2,"tier_adm_region"],
    ["tier_adm_village","org_preset_admin","乡",1,"tier_adm_county"],
    # Commerce
    ["tier_com_hq","org_preset_commerce","总行",5,""],
    ["tier_com_branch","org_preset_commerce","分行",4,"tier_com_hq"],
    ["tier_com_house","org_preset_commerce","商号",3,"tier_com_branch"],
    ["tier_com_shop","org_preset_commerce","商铺",2,"tier_com_house"],
    ["tier_com_vendor","org_preset_commerce","货郎",1,"tier_com_shop"],
]

r = ws3.max_row + 1
for row_data in new_presets:
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=val).font = F()
    r += 1

wb3.save(path); wb3.close(); print("编制预设 (updated) done")

# =============================================
# 战斗战术.xlsx: tactical presets
# =============================================
wb4 = openpyxl.Workbook()
ws = wb4.active; ws.title = "tactics"
ws.append(["id","name_zh","description","cover_priority","flank_enabled","retreat_threshold","fire_discipline","autonomy"])
ws.append(["唯一ID","战术名","描述","掩体优先级","侧翼","撤退阈值","火力纪律","自主度"])
for c in range(1,9): ws.cell(row=1,column=c).font=F(True); ws.cell(row=2,column=c).font=F()
tactics = [
    ["tac_aggressive","猛攻","主动冲击不惜代价","low","yes","0.2","loose","high"],
    ["tac_defensive","坚守","固守阵地寸步不退","high","no","0.5","tight","low"],
    ["tac_guerrilla","游击","打了就跑骚扰为主","low","yes","0.1","loose","high"],
    ["tac_balanced","均衡","灵活应对战场局势","medium","yes","0.35","medium","medium"],
    ["tac_sniper","狙击","远程优先避近战","high","no","0.3","tight","low"],
]
for row in tactics: ws.append(row)
for c in range(1,9): ws.column_dimensions[chr(64+c)].width = 18

wb4.save(f'{dir_}/战斗战术.xlsx'); wb4.close(); print("战斗战术.xlsx done")

# Summary
files = [f for f in os.listdir(dir_) if f.endswith('.xlsx') and not f.startswith('~$')]
print(f"\n总计 {len(files)} 个 xlsx:")
for f in sorted(files):
    wb = openpyxl.load_workbook(f'{dir_}/{f}')
    sheets = ', '.join(wb.sheetnames)
    print(f"  {f} -> [{sheets}]")
    wb.close()
