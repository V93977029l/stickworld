"""
Excel 导出管线: config/excel/*.xlsx -> config/**/*.tres
用法: python tools/export_excel.py [--dry-run]
依赖: pip install openpyxl
"""
import openpyxl, os, sys, json, re

EXCEL_DIR = "stick-world/config/excel"
CONFIG_DIR = "stick-world/config"
DRY_RUN = "--dry-run" in sys.argv

def fix_type(val):
    """自动检测并转换单元格值为正确类型"""
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        s = val.strip()
        if s.lower() == "true": return True
        if s.lower() == "false": return False
        if s.lower() == "none" or s == "": return None
        try: return int(s)
        except: pass
        try: return float(s)
        except: pass
        if s.startswith("[") and s.endswith("]"):
            try: return json.loads(s)
            except: pass
        if s.startswith("{") and s.endswith("}"):
            try: return json.loads(s)
            except: pass
        return s
    return val

def make_tres_content(type_path, rows):
    """生成 Godot .tres 文件内容"""
    lines = [
        '[gd_resource type="Resource" script_class="BalanceResource" load_steps=2 format=3 uid="uid://{0}"]'.format(
            os.urandom(8).hex()[:12]
        ),
        '',
        '[ext_resource type="Script" path="res://config/balance/balance_resource.gd" id="1"]',
        '',
        '[resource]',
        'script = ExtResource("1")',
    ]
    # variables dict with "data" key for rows
    lines.append('variables = {')
    lines.append('"data": [')
    for i, row in enumerate(rows):
        parts = []
        for k, v in row.items():
            if v is None:
                continue
            if isinstance(v, bool):
                parts.append(f'"{k}": {str(v).lower()}')
            elif isinstance(v, (int, float)):
                parts.append(f'"{k}": {v}')
            elif isinstance(v, str):
                escaped = v.replace('\\', '\\\\').replace('"', '\\"')
                parts.append(f'"{k}": "{escaped}"')
            elif isinstance(v, list):
                parts.append(f'"{k}": {json.dumps(v)}')
            elif isinstance(v, dict):
                parts.append(f'"{k}": {json.dumps(v)}')
        comma = "," if i < len(rows) - 1 else ""
        lines.append("    {" + ", ".join(parts) + "}" + comma)
    lines.append("]")
    lines.append("}")
    return "\n".join(lines) + "\n"

def sheet_to_tres_name(sheet_name):
    """将中文 Sheet 名映射到英文 .tres 文件名"""
    name_map = {
        "火柴人": "stickmen", "武器": "weapons", "盔甲": "armors",
        "资源": "resources",
        "建筑": "buildings",
        "科技": "techs",
        "平衡变量": "variables",
        "编制预设": "presets",
        "地块": "regions", "势力": "factions", "文明差异": "civ_differences",
        "运输载体": "carriers", "道路": "roads",
        "战术": "tactics",
    }
    return name_map.get(sheet_name, sheet_name.lower().replace(" ", "_"))

def sheet_to_output_dir(filename, sheet_name):
    """将 Excel 文件名映射到 config 输出子目录"""
    # 如果 Sheet 的 A1 或文件名本身定义了输出目录，用它
    name_map = {
        "单位数据": "units",
        "资源数据": "resources",
        "建筑数据": "buildings",
        "科技树": "tech",
        "平衡变量": "balance",
        "编制预设": "formations",
        "扩张数据": "expansion",
        "运输数据": "logistics",
        "战斗战术": "combat",
    }
    base = os.path.splitext(filename)[0]
    return name_map.get(base, base.lower())

def process_excel(filepath, filename):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    errors = []
    generated = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 3:
            continue  # 至少要有表头+中文说明+1行数据

        # R1 = 英文字段名, R2 = 中文说明, R3+ = 数据
        headers = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=c).value
            if val:
                # 去掉 * 标记（必填列）
                clean = str(val).replace("*", "").strip()
                if clean and clean not in ("None", ""):
                    headers.append((c, clean))
                else:
                    # 空列名 → 停止(后面的列是注释用)
                    break

        if not headers:
            continue

        # 解析数据行
        rows = []
        for ri in range(3, ws.max_row + 1):
            row_dict = {}
            for col_idx, field_name in headers:
                cell = ws.cell(row=ri, column=col_idx)
                val = fix_type(cell.value)
                if val is not None:
                    row_dict[field_name] = val
            if row_dict:
                rows.append(row_dict)

        if not rows:
            continue

        # 验证
        id_set = set()
        for i, row in enumerate(rows):
            rid = row.get("id", f"<row_{i+3}>")
            if rid in id_set:
                errors.append(f"{filename}/{sheet_name}: 重复 id '{rid}' (第{i+3}行)")
            id_set.add(rid)

        if errors:
            print(f"  ⚠️ {filename}/{sheet_name}: 验证失败")
            for e in errors:
                print(f"     {e}")
            continue

        # 生成 .tres
        output_dir = sheet_to_output_dir(filename, sheet_name)
        full_dir = os.path.join(CONFIG_DIR, output_dir)
        if not DRY_RUN:
            os.makedirs(full_dir, exist_ok=True)

        tres_name = sheet_to_tres_name(sheet_name) + ".tres"
        tres_path = os.path.join(full_dir, tres_name)
        content = make_tres_content(f"{output_dir}.{sheet_name}", rows)

        if DRY_RUN:
            print(f"  ✅ {filename}/{sheet_name} -> config/{output_dir}/{tres_name} ({len(rows)} rows) [DRY RUN]")
        else:
            with open(tres_path, "w", encoding="utf-8") as f:
                f.write(content)
            print(f"  ✅ {filename}/{sheet_name} -> config/{output_dir}/{tres_name} ({len(rows)} rows)")

        generated.append((output_dir, sheet_name, len(rows)))

    wb.close()
    return generated, errors

def main():
    if not os.path.isdir(EXCEL_DIR):
        print(f"错误: Excel 目录不存在: {EXCEL_DIR}")
        sys.exit(1)

    files = sorted([f for f in os.listdir(EXCEL_DIR) if f.endswith(".xlsx") and not f.startswith("~$")])
    if not files:
        print("没有找到 .xlsx 文件")
        return

    total_generated = 0
    total_errors = 0

    for f in files:
        path = os.path.join(EXCEL_DIR, f)
        try:
            generated, errors = process_excel(path, f)
            total_generated += len(generated)
            total_errors += len(errors)
        except Exception as e:
            print(f"  ❌ {f}: {e}")
            total_errors += 1

    print(f"\n导出完成: {total_generated} 个 .tres 生成, {total_errors} 个错误")

if __name__ == "__main__":
    main()
