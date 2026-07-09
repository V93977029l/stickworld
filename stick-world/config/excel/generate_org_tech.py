"""生成 编制预设.xlsx 和 科技树.xlsx"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
from zipfile import ZipFile, ZIP_DEFLATED
import shutil
import tempfile

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ── 样式（与 Agent E2 格式一致） ──
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
DATA_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")


def apply_header_style(ws, col_count):
    for row in [1, 2]:
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER


def apply_data_style(ws, start_row, end_row, col_count):
    for row in range(start_row, end_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.alignment = LEFT_ALIGN
            cell.border = THIN_BORDER


def auto_width(ws, col_count, min_width=10, max_width=50):
    for col in range(1, col_count + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            for cell_val in row:
                if cell_val is not None:
                    s = str(cell_val)
                    length = sum(2 if ord(c) > 127 else 1 for c in s)
                    max_len = max(max_len, length)
        ws.column_dimensions[letter].width = max(min_width, min(max_len + 2, max_width))


def write_data_row(ws, row_idx, row_data):
    """写入一行数据，空字符串用 inlineStr 类型标记"""
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if val == "":
            cell.data_type = 's'
            cell.value = ""
        else:
            cell.value = val


def fix_empty_inline_str(xlsx_path):
    """
    修复 openpyxl 生成的空 inlineStr 单元格。
    openpyxl 输出 <c r="E3" t="inlineStr" /> （缺少 <is> 子元素），
    修复为 <c r="E3" t="inlineStr"><is><t></t></is></c>。
    """
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)

    with ZipFile(xlsx_path, 'r') as zin:
        with ZipFile(tmp_path, 'w', ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith('xl/worksheets/sheet') and item.filename.endswith('.xml'):
                    text = data.decode('utf-8')
                    # 修复：<c ... t="inlineStr" /> → <c ... t="inlineStr"><is><t></t></is></c>
                    text = re.sub(
                        r'(<c[^>]*\st="inlineStr"[^>]*)(/>)',
                        r'\1><is><t></t></is></c>',
                        text,
                    )
                    data = text.encode('utf-8')
                zout.writestr(item, data)

    shutil.move(tmp_path, xlsx_path)


# ═══════════════════════════════════════════════
#  表 A: 编制预设.xlsx → Sheet "presets"
# ═══════════════════════════════════════════════
def create_organizations():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "presets"

    en_headers = ["id", "name_zh", "tag", "preset_data_json"]
    zh_headers = ["ID", "名称", "标签", "预设数据JSON"]

    for col_idx, (en, zh) in enumerate(zip(en_headers, zh_headers), 1):
        ws.cell(row=1, column=col_idx, value=en)
        ws.cell(row=2, column=col_idx, value=zh)

    data = [
        [
            "org_preset_military",
            "标准军事编制",
            "military",
            '{"tiers":[{"name":"师","level":4,"children":[{"name":"团","level":3,"subdivisions":3,"children":[{"name":"连","level":2,"subdivisions":4,"children":[{"name":"排","level":1,"subdivisions":4,"children":[]}]}]}]}]}',
        ],
        [
            "org_preset_academy",
            "科学院架构",
            "research",
            '{"tiers":[{"name":"科学院","level":5,"children":[{"name":"研究所","level":4,"subdivisions":5,"children":[{"name":"研究室","level":3,"subdivisions":4,"children":[{"name":"课题组","level":2,"subdivisions":3,"children":[]}]}]}]}]}',
        ],
    ]

    for row_idx, row_data in enumerate(data, 3):
        write_data_row(ws, row_idx, row_data)

    col_count = len(en_headers)
    apply_header_style(ws, col_count)
    apply_data_style(ws, 3, 2 + len(data), col_count)
    auto_width(ws, col_count, max_width=100)
    ws.freeze_panes = "A3"

    path = os.path.join(OUTPUT_DIR, "编制预设.xlsx")
    wb.save(path)
    fix_empty_inline_str(path)
    print(f"编制预设.xlsx 已生成 ({len(data)} 行数据)")


# ═══════════════════════════════════════════════
#  表 B: 科技树.xlsx → Sheet "techs"
# ═══════════════════════════════════════════════
def create_tech_tree():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "techs"

    en_headers = ["id", "name_zh", "tier", "category", "prerequisites", "research_cost", "unlocks"]
    zh_headers = ["ID", "名称", "层级", "类别", "前置科技", "研究消耗", "解锁内容"]

    for col_idx, (en, zh) in enumerate(zip(en_headers, zh_headers), 1):
        ws.cell(row=1, column=col_idx, value=en)
        ws.cell(row=2, column=col_idx, value=zh)

    data = [
        # ── Tier 1 ──
        ["tech_stone_tools",              "石器工具",   1, "economy",        "",                         100, '["stone_axe","stone_pickaxe","basic_quarry"]'],
        ["tech_basic_combat",             "基础战斗",   1, "military",       "",                         120, '["warrior","spear","wooden_shield"]'],
        ["tech_primitive_administration", "原始行政管理", 1, "administration", "",                          80, '["tribal_council","basic_laws","census_tent"]'],
        # ── Tier 2 ──
        ["tech_bronze_working",   "青铜冶炼", 2, "economy",        "tech_stone_tools",                          300, '["bronze_axe","bronze_pickaxe","smelter","copper_mine"]'],
        ["tech_formation_tactics","阵型战术", 2, "military",       "tech_basic_combat",                         350, '["spearman","shield_wall","barracks","military_drill"]'],
        ["tech_early_writing",    "早期文字", 2, "administration", "tech_primitive_administration",             250, '["scribe","clay_tablet","record_keeping","royal_decree"]'],
        ["tech_herbal_medicine",  "草药医学", 2, "science",        "tech_primitive_administration",             280, '["herbalist","field_hospital","healing_salve"]'],
        # ── Tier 3 ──
        ["tech_iron_working",    "铁器锻造", 3, "economy",        "tech_bronze_working",                       600, '["iron_axe","iron_pickaxe","blast_furnace","iron_mine"]'],
        ["tech_cavalry_warfare", "骑兵战术", 3, "military",       "tech_formation_tactics,tech_bronze_working", 700, '["horseman","cavalry_archer","stable","horse_breeding"]'],
        ["tech_philosophy",      "哲学思想", 3, "science",        "tech_early_writing,tech_herbal_medicine",    550, '["philosopher","academy","natural_philosophy","logic"]'],
        ["tech_bureaucracy",     "官僚制度", 3, "administration", "tech_early_writing",                        500, '["bureaucrat","tax_office","census","archives"]'],
        ["tech_trade_networks",  "贸易网络", 3, "economy",        "tech_early_writing,tech_bronze_working",    650, '["merchant","marketplace","trade_route","currency"]'],
    ]

    for row_idx, row_data in enumerate(data, 3):
        write_data_row(ws, row_idx, row_data)

    col_count = len(en_headers)
    apply_header_style(ws, col_count)
    apply_data_style(ws, 3, 2 + len(data), col_count)
    auto_width(ws, col_count, max_width=80)
    ws.freeze_panes = "A3"

    path = os.path.join(OUTPUT_DIR, "科技树.xlsx")
    wb.save(path)
    fix_empty_inline_str(path)
    print(f"科技树.xlsx 已生成 ({len(data)} 行数据)")


# ═══════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════
if __name__ == "__main__":
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    create_organizations()
    create_tech_tree()
    print("全部完成！")