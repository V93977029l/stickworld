"""生成策划数据 Excel 表（资源数据 / 建筑数据 / 平衡变量）"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ── 样式定义 ──
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # 灰色背景
DATA_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")


def apply_header_style(ws, col_count):
    """给前两行（英文+中文表头）加粗+灰底+边框+居中"""
    for row in [1, 2]:
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER


def apply_data_style(ws, start_row, end_row, col_count):
    """给数据行加边框+字体"""
    for row in range(start_row, end_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER


def auto_width(ws, col_count, min_width=10, max_width=40):
    """自动列宽"""
    for col in range(1, col_count + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            for cell_val in row:
                if cell_val is not None:
                    # 中文字符算2个宽度
                    s = str(cell_val)
                    length = sum(2 if ord(c) > 127 else 1 for c in s)
                    max_len = max(max_len, length)
        ws.column_dimensions[letter].width = max(min_width, min(max_len + 2, max_width))


# ═══════════════════════════════════════════════
#  表 A: 资源数据.xlsx
# ═══════════════════════════════════════════════
def create_resources():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "resources"

    # 英文表头（第1行）
    en_headers = ["id", "name_zh", "category", "base_price", "weight_per_unit",
                  "perishable", "icon_path"]
    # 中文表头（第2行）
    zh_headers = ["ID", "名称", "类别", "基础价格", "单位重量",
                  "是否易腐", "图标路径"]

    for col_idx, (en, zh) in enumerate(zip(en_headers, zh_headers), 1):
        ws.cell(row=1, column=col_idx, value=en)
        ws.cell(row=2, column=col_idx, value=zh)

    # 数据（第3行起）
    data = [
        ["res_food",          "食物",     "basic",    5,  1.0, True,  "res://assets/icons/food.png"],
        ["res_wood",          "木材",     "basic",    3,  2.0, False, "res://assets/icons/wood.png"],
        ["res_stone",         "石料",     "basic",    4,  5.0, False, "res://assets/icons/stone.png"],
        ["res_metal_ore",     "金属矿",   "basic",    8,  8.0, False, "res://assets/icons/metal_ore.png"],
        ["res_black_asphalt", "黑色沥青", "strategic", 15, 6.0, False, "res://assets/icons/black_asphalt.png"],
        ["res_iron_ingot",    "铁锭",     "processed", 20, 7.0, False, "res://assets/icons/iron_ingot.png"],
        ["res_silk",          "丝绸",     "luxury",    50, 0.5, False, "res://assets/icons/silk.png"],
    ]

    for row_idx, row_data in enumerate(data, 3):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    col_count = len(en_headers)
    apply_header_style(ws, col_count)
    apply_data_style(ws, 3, 2 + len(data), col_count)
    auto_width(ws, col_count)

    # 冻结表头
    ws.freeze_panes = "A3"

    path = os.path.join(OUTPUT_DIR, "资源数据.xlsx")
    wb.save(path)
    print(f"✓ 已生成: {path} ({len(data)} 行数据)")


# ═══════════════════════════════════════════════
#  表 B: 建筑数据.xlsx
# ═══════════════════════════════════════════════
def create_buildings():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "buildings"

    en_headers = [
        "id", "name_zh", "type", "tier", "build_time",
        "build_cost_food", "build_cost_wood", "build_cost_stone", "build_cost_metal",
        "max_hp", "workers_required", "unlocked_by_tech", "description"
    ]
    zh_headers = [
        "ID", "名称", "类型", "等级", "建造时间",
        "食物消耗", "木材消耗", "石料消耗", "金属消耗",
        "最大生命值", "所需工人", "解锁科技", "描述"
    ]

    for col_idx, (en, zh) in enumerate(zip(en_headers, zh_headers), 1):
        ws.cell(row=1, column=col_idx, value=en)
        ws.cell(row=2, column=col_idx, value=zh)

    data = [
        ["bld_house",    "民居", "house",    1, 30, 50,  100, 50,  0,  200, 0, "none",              "提供人口居住空间"],
        ["bld_farm",     "农场", "farm",     1, 45, 20,  80,  30,  0,  150, 2, "tech_agriculture",  "生产食物资源"],
        ["bld_workshop", "工坊", "workshop",  2, 60, 30,  120, 60,  20, 300, 3, "tech_crafting",     "加工基础资源为高级资源"],
        ["bld_barracks", "兵营", "barracks",  2, 90, 80,  150, 100, 50, 500, 5, "tech_military_1",   "训练战斗单位"],
        ["bld_market",   "集市", "market",    2, 50, 100, 80,  40,  10, 250, 2, "tech_trade",        "进行资源交易"],
        ["bld_academy",  "学院", "academy",   3, 120, 150, 200, 150, 80, 600, 4, "tech_education",   "研究科技"],
    ]

    for row_idx, row_data in enumerate(data, 3):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    col_count = len(en_headers)
    apply_header_style(ws, col_count)
    apply_data_style(ws, 3, 2 + len(data), col_count)
    auto_width(ws, col_count)
    ws.freeze_panes = "A3"

    path = os.path.join(OUTPUT_DIR, "建筑数据.xlsx")
    wb.save(path)
    print(f"✓ 已生成: {path} ({len(data)} 行数据)")


# ═══════════════════════════════════════════════
#  表 C: 平衡变量.xlsx
# ═══════════════════════════════════════════════
def create_balance():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "variables"

    en_headers = ["id", "category", "value", "min", "max", "step", "description"]
    zh_headers = ["ID", "类别", "当前值", "最小值", "最大值", "步长", "描述"]

    for col_idx, (en, zh) in enumerate(zip(en_headers, zh_headers), 1):
        ws.cell(row=1, column=col_idx, value=en)
        ws.cell(row=2, column=col_idx, value=zh)

    data = [
        # combat
        ["var_attack_base",        "combat",      10,   1,   100,   1,  "基础攻击力"],
        ["var_defense_base",       "combat",      5,    1,   100,   1,  "基础防御力"],
        # economy
        ["var_tax_rate",           "economy",     0.15, 0.0, 0.5,   0.01, "基础税率"],
        ["var_trade_efficiency",   "economy",     1.0,  0.1, 5.0,   0.1,  "贸易效率倍率"],
        # tech
        ["var_research_speed",     "tech",        1.0,  0.1, 10.0,  0.1,  "科技研究速度倍率"],
        ["var_tech_cost_scale",    "tech",        100,  10,  1000,  10,   "科技研究基础消耗"],
        # org
        ["var_org_maintenance",    "org",         10,   1,   100,   1,  "组织每回合维护费"],
        ["var_org_loyalty_decay",  "org",         0.5,  0.0, 5.0,   0.1,  "组织忠诚度每回合衰减"],
        # expansion
        ["var_expand_cost_base",   "expansion",   200,  50,  1000,  10,   "扩张基础消耗"],
        ["var_expand_cooldown",    "expansion",   5,    1,   30,    1,  "扩张冷却回合数"],
        # logistics
        ["var_supply_range",       "logistics",   10,   1,   50,    1,  "补给线最大范围"],
        ["var_transport_speed",    "logistics",   1.0,  0.1, 5.0,   0.1,  "运输速度倍率"],
        # population
        ["var_pop_growth_rate",    "population",  0.02, 0.0, 0.1,   0.001, "人口增长率"],
        ["var_pop_happiness_base", "population",  50,   0,   100,   1,  "基础幸福度"],
        # global
        ["var_game_speed",         "global",      1.0,  0.25, 4.0,  0.25, "游戏速度倍率"],
        ["var_day_duration",       "global",      10,   1,   60,    1,  "每回合代表天数"],
    ]

    for row_idx, row_data in enumerate(data, 3):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    col_count = len(en_headers)
    apply_header_style(ws, col_count)
    apply_data_style(ws, 3, 2 + len(data), col_count)
    auto_width(ws, col_count)
    ws.freeze_panes = "A3"

    # 按 category 添加数据验证（可选——方便策划填写时约束）
    # 这里不做强制，保持灵活性

    path = os.path.join(OUTPUT_DIR, "平衡变量.xlsx")
    wb.save(path)
    print(f"✓ 已生成: {path} ({len(data)} 行数据)")


# ═══════════════════════════════════════════════
#  Main
# ═══════════════════════════════════════════════
if __name__ == "__main__":
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    create_resources()
    create_buildings()
    create_balance()
    print("\n全部完成！")