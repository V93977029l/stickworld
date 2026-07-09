#!/usr/bin/env python3
"""
export_excel.py —— 将 config/excel/*.xlsx 导出为 Godot .tres 资源文件。

用法:
    python tools/export_excel.py              # 导出所有
    python tools/export_excel.py --dry-run    # 只校验不导出

Excel 格式约定:
    - 第 1 行: 英文字段名（.tres 的键）
    - 第 2 行: 中文字段说明（跳过）
    - 第 3 行起: 数据行

输出:
    config/excel/<文件名>.xlsx → config/<文件名>/<sheet名>.tres
"""

import json
import re
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------

# 项目根目录（相对于本脚本位置）
PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_DIR = PROJECT_ROOT / "config" / "excel"
CONFIG_DIR = PROJECT_ROOT / "config"
ASSETS_DIR = PROJECT_ROOT / "assets"

# Godot .tres 模板使用的脚本路径
BALANCE_RESOURCE_SCRIPT = "res://config/balance/balance_resource.gd"

# 必填列标记：列名以 * 结尾
REQUIRED_COLUMN_SUFFIX = "*"

# 引用列模式：xxx_id 格式
REF_COLUMN_PATTERN = re.compile(r"^(.+)_id$")


# ---------------------------------------------------------------------------
# 类型检测与转换
# ---------------------------------------------------------------------------

def detect_and_convert(value):
    """自动检测单元格值类型并转换为对应的 Python 原生类型。

    检测顺序:
        1. None / 空字符串 → None
        2. 已为 bool → 原样返回
        3. 已为 int / float → 原样返回
        4. 字符串 "true" / "false" → bool
        5. 字符串整数 → int
        6. 字符串浮点数 → float
        7. 以 [ 开头 ] 结尾 → Array（JSON 解析）
        8. 以 { 开头 } 结尾 → Dict（JSON 解析）
        9. 其他 → string
    """
    if value is None:
        return None

    # openpyxl 可能已经返回了 Python 原生类型
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value

    s = str(value).strip()
    if s == "":
        return None

    # bool
    if s.lower() == "true":
        return True
    if s.lower() == "false":
        return False

    # int
    try:
        if "." not in s and "e" not in s.lower():
            return int(s)
    except ValueError:
        pass

    # float
    try:
        return float(s)
    except ValueError:
        pass

    # Array: [ ... ]
    if s.startswith("[") and s.endswith("]"):
        try:
            parsed = json.loads(s)
            if isinstance(parsed, list):
                return parsed
        except json.JSONDecodeError:
            pass

    # Dict: { ... }
    if s.startswith("{") and s.endswith("}"):
        try:
            parsed = json.loads(s)
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            pass

    # 默认：字符串
    return s


# ---------------------------------------------------------------------------
# Godot .tres 格式化
# ---------------------------------------------------------------------------

def to_godot_value(value, indent=0):
    """将 Python 值转换为 Godot .tres 文本格式的字符串。"""
    prefix = "\t" * indent

    if value is None:
        return "null"

    if isinstance(value, bool):
        return "true" if value else "false"

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        # 确保浮点数有小数点
        s = repr(value)
        if "." not in s:
            s += ".0"
        return s

    if isinstance(value, str):
        # 转义引号和反斜杠
        escaped = value.replace("\\", "\\\\").replace('"', '\\"')
        return f'"{escaped}"'

    if isinstance(value, list):
        if not value:
            return "[]"
        items = [to_godot_value(v, indent + 1) for v in value]
        # 单行短数组
        oneline = ", ".join(items)
        if len(oneline) <= 80:
            return f"[{oneline}]"
        # 多行数组
        inner = ",\n".join(f"{prefix}\t{item}" for item in items)
        return f"[\n{prefix}\t{inner}\n{prefix}]"

    if isinstance(value, dict):
        if not value:
            return "{}"
        items = []
        for k, v in value.items():
            v_str = to_godot_value(v, indent + 1)
            items.append(f'"{k}": {v_str}')
        oneline = ", ".join(items)
        if len(oneline) <= 80:
            return f"{{{oneline}}}"
        inner = ",\n".join(f"{prefix}\t{item}" for item in items)
        return f"{{\n{prefix}\t{inner}\n{prefix}}}"

    # 兜底
    return f'"{str(value)}"'


def format_dict_as_godot(data, indent=0):
    """将一个 Python dict 格式化为 Godot 字典的多行文本。"""
    prefix = "\t" * indent
    if not data:
        return "{}"
    lines = []
    for key, value in data.items():
        v_str = to_godot_value(value, indent + 1)
        lines.append(f'{prefix}\t"{key}": {v_str}')
    inner = ",\n".join(lines)
    return f"{{\n{inner}\n{prefix}}}"


# ---------------------------------------------------------------------------
# .tres 文件生成
# ---------------------------------------------------------------------------

def generate_tres(resource_name, data_rows, source_file, source_sheet, description=""):
    """生成 .tres 文件内容字符串。

    使用现有 balance_resource.gd 作为基类，将数据存入 variables.data。
    """
    header = f"""[gd_resource type="Resource" load_steps=2 format=3]

[ext_resource type="Script" path="{BALANCE_RESOURCE_SCRIPT}" id="1"]

[resource]
script = ExtResource("1")
resource_name = "{resource_name}"
_meta = {{
\t"resource_name": "{resource_name}",
\t"version": "0.1.0",
\t"description": "{description}",
\t"category": "excel_export",
\t"source_file": "{source_file}",
\t"source_sheet": "{source_sheet}"
}}"""

    # 格式化数据数组
    if not data_rows:
        data_block = "[]"
    else:
        row_strings = []
        for row in data_rows:
            row_str = format_dict_as_godot(row, indent=2)
            row_strings.append(f"\t\t{row_str}")
        data_inner = ",\n".join(row_strings)
        data_block = f"[\n{data_inner}\n\t]"

    body = f"""
variables = {{
\t"data": {data_block}
}}"""

    return header + body + "\n"


# ---------------------------------------------------------------------------
# Excel 读取
# ---------------------------------------------------------------------------

def parse_sheet_with_raw(ws):
    """解析工作表，同时返回原始 headers（含 * 标记）和清洗后的 headers。

    约定:
        - 第 1 行: 英文字段名（可带 * 表示必填）
        - 第 2 行: 中文说明（跳过）
        - 第 3 行起: 数据

    返回: (raw_headers, clean_headers, data_rows, errors)
    """
    errors = []
    rows = list(ws.iter_rows(min_row=1, values_only=True))

    if len(rows) < 3:
        errors.append("数据不足（至少需要 3 行：表头 + 说明 + 数据）")
        return [], [], [], errors

    # 原始字段名（含 * 标记）
    raw_headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    # 清洗后的字段名（去掉 *）
    clean_headers = [h.rstrip(REQUIRED_COLUMN_SUFFIX) for h in raw_headers]

    # 验证字段名
    seen = set()
    for h in clean_headers:
        if not h:
            continue
        if h in seen:
            errors.append(f"字段名重复: '{h}'")
        seen.add(h)

    if errors:
        return raw_headers, clean_headers, [], errors

    # 数据行
    data_rows = []
    for row_idx, row in enumerate(rows[2:], start=3):
        row_data = {}
        row_has_data = False
        for header, value in zip(clean_headers, row):
            if not header:
                continue
            converted = detect_and_convert(value)
            row_data[header] = converted
            if converted is not None:
                row_has_data = True

        if row_has_data:
            data_rows.append(row_data)

    return raw_headers, clean_headers, data_rows, errors


# ---------------------------------------------------------------------------
# 验证
# ---------------------------------------------------------------------------

def validate_sheet(sheet_name, file_name, headers, data_rows, all_ids_map):
    """验证单个 Sheet 的数据。

    检查项:
        1. id 列不可重复
        2. 必填列（带 * 标记）不能为空
        3. 引用完整性（xxx_id 列必须指向对应 sheet 的 id）

    all_ids_map: {sheet_name: set of id values}
    返回错误列表。
    """
    errors = []

    # 检查 id 列
    if "id" in headers:
        seen_ids = set()
        for row_idx, row in enumerate(data_rows, start=3):
            row_id = row.get("id")
            if row_id is None:
                errors.append(
                    f"[{file_name}] Sheet '{sheet_name}' 第 {row_idx} 行: id 为空"
                )
            elif row_id in seen_ids:
                errors.append(
                    f"[{file_name}] Sheet '{sheet_name}' 第 {row_idx} 行: id '{row_id}' 重复"
                )
            else:
                seen_ids.add(row_id)

    # 检查引用完整性
    for header in headers:
        m = REF_COLUMN_PATTERN.match(header)
        if not m:
            continue
        ref_sheet = m.group(1)  # 如 weapon_id → weapon

        # 尝试匹配目标 sheet（支持单复数变体）
        target_ids = None
        for candidate in [ref_sheet, ref_sheet + "s"]:
            if candidate in all_ids_map:
                target_ids = all_ids_map[candidate]
                break

        if target_ids is None:
            # 目标 sheet 不存在，跳过引用检查
            continue

        for row_idx, row in enumerate(data_rows, start=3):
            ref_val = row.get(header)
            if ref_val is None:
                continue
            ref_str = str(ref_val)
            if ref_str not in target_ids:
                errors.append(
                    f"[{file_name}] Sheet '{sheet_name}' 第 {row_idx} 行: "
                    f"'{header}' = '{ref_str}' 指向的 id 在 sheet '{ref_sheet}' 中不存在"
                )

    return errors


def validate_required_columns(file_name, sheet_name, raw_headers, data_rows):
    """检查必填列（带 * 标记的列）不能为空。"""
    errors = []
    required_indices = []
    for idx, h in enumerate(raw_headers):
        if h.endswith(REQUIRED_COLUMN_SUFFIX):
            clean = h.rstrip(REQUIRED_COLUMN_SUFFIX)
            required_indices.append((idx, clean))

    for row_idx, row in enumerate(data_rows, start=3):
        for _, col_name in required_indices:
            val = row.get(col_name)
            if val is None or (isinstance(val, str) and val.strip() == ""):
                errors.append(
                    f"[{file_name}] Sheet '{sheet_name}' 第 {row_idx} 行: "
                    f"必填列 '{col_name}' 为空"
                )

    return errors


# ---------------------------------------------------------------------------
# 图片提取（可选，占位）
# ---------------------------------------------------------------------------

def extract_images(ws, sheet_name, output_asset_dir):
    """从工作表中提取嵌入的图片，保存到 assets/ 对应目录。

    返回: {图片位置(row, col): 资源路径} 的映射字典。
    注意: 图片提取依赖 openpyxl 的图片 API，需要先实现结构。
    """
    image_map = {}
    if not hasattr(ws, "_images"):
        return image_map

    # 确保目标目录存在
    output_asset_dir.mkdir(parents=True, exist_ok=True)

    for img in ws._images:
        # 获取图片锚点位置
        if hasattr(img, "anchor") and hasattr(img.anchor, "_from"):
            row = img.anchor._from.row + 1  # 1-based
            col = img.anchor._from.col + 1  # 1-based
            # 生成文件名
            img_ext = img.format if hasattr(img, "format") else "png"
            img_name = f"{sheet_name}_r{row}_c{col}.{img_ext.lower()}"
            img_path = output_asset_dir / img_name
            # 保存图片数据
            with open(img_path, "wb") as f:
                f.write(img._data())
            image_map[(row, col)] = f"res://assets/{sheet_name}/{img_name}"

    return image_map


# ---------------------------------------------------------------------------
# 主处理流程
# ---------------------------------------------------------------------------

def find_excel_files():
    """扫描 config/excel/ 下所有 .xlsx 文件。"""
    if not EXCEL_DIR.exists():
        print(f"[警告] Excel 目录不存在: {EXCEL_DIR}")
        return []
    return sorted(EXCEL_DIR.glob("*.xlsx"))


def export_all(dry_run=False):
    """主入口：扫描、解析、验证、导出。"""
    import openpyxl

    excel_files = find_excel_files()
    if not excel_files:
        print("没有找到 .xlsx 文件，跳过。")
        return True

    all_errors = []
    all_parsed = {}  # {(file_name, sheet_name): (headers, data_rows, raw_headers)}
    all_ids_map = {}  # {sheet_name: set of ids}

    # ── 第一遍：解析所有 Sheet ──
    print("=" * 60)
    print("第一遍：解析所有 Excel 文件...")
    print("=" * 60)

    for xlsx_path in excel_files:
        file_name = xlsx_path.name
        print(f"\n📄 {file_name}")

        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        except Exception as e:
            all_errors.append(f"[{file_name}] 无法打开文件: {e}")
            continue

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            raw_headers, headers, data_rows, sheet_errors = parse_sheet_with_raw(ws)

            if sheet_errors:
                for err in sheet_errors:
                    all_errors.append(f"[{file_name}] Sheet '{sheet_name}': {err}")
                continue

            key = (file_name, sheet_name)
            all_parsed[key] = (headers, data_rows, raw_headers)

            # 收集 id 用于引用完整性检查
            if "id" in headers:
                ids = set()
                for row in data_rows:
                    row_id = row.get("id")
                    if row_id is not None:
                        ids.add(str(row_id))
                all_ids_map[sheet_name] = ids

            print(f"  ├─ Sheet '{sheet_name}': {len(data_rows)} 行数据, {len(headers)} 列")

        wb.close()

    # ── 第二遍：验证 ──
    print(f"\n{'=' * 60}")
    print("第二遍：数据验证...")
    print("=" * 60)

    for (file_name, sheet_name), (headers, data_rows, raw_headers) in all_parsed.items():
        # 必填列检查
        req_errors = validate_required_columns(file_name, sheet_name, raw_headers, data_rows)
        all_errors.extend(req_errors)

        # id 唯一性 + 引用完整性
        ref_errors = validate_sheet(sheet_name, file_name, headers, data_rows, all_ids_map)
        all_errors.extend(ref_errors)

    # ── 报告验证结果 ──
    if all_errors:
        print(f"\n❌ 发现 {len(all_errors)} 个错误:")
        for err in all_errors:
            print(f"  {err}")

    if dry_run:
        print(f"\n{'=' * 60}")
        if all_errors:
            print(f"❌ 干跑模式：校验失败 ({len(all_errors)} 个错误)")
        else:
            print("✅ 干跑模式：校验通过")
        print("=" * 60)
        return len(all_errors) == 0

    if all_errors:
        print(f"\n⚠️  存在 {len(all_errors)} 个验证错误，是否继续导出？(y/n): ", end="")
        choice = input().strip().lower()
        if choice != "y":
            print("已取消导出。")
            return False

    # ── 第三遍：导出 .tres ──
    print(f"\n{'=' * 60}")
    print("第三遍：导出 .tres 文件...")
    print("=" * 60)

    exported_count = 0
    for (file_name, sheet_name), (headers, data_rows, _) in all_parsed.items():
        # 输出目录: config/<excel文件名>/
        base_name = Path(file_name).stem  # 去掉 .xlsx
        output_dir = CONFIG_DIR / base_name
        output_dir.mkdir(parents=True, exist_ok=True)

        output_path = output_dir / f"{sheet_name}.tres"

        # 生成 .tres 内容
        description = f"从 config/excel/{file_name} → {sheet_name} sheet 自动生成"
        tres_content = generate_tres(
            resource_name=sheet_name,
            data_rows=data_rows,
            source_file=f"config/excel/{file_name}",
            source_sheet=sheet_name,
            description=description,
        )

        with open(output_path, "w", encoding="utf-8") as f:
            f.write(tres_content)

        exported_count += 1
        print(f"  ✅ {output_path.relative_to(PROJECT_ROOT)}  ({len(data_rows)} 行)")

    # ── 总结 ──
    print(f"\n{'=' * 60}")
    print(f"✅ 导出完成: {exported_count} 个 .tres 文件")
    if all_errors:
        print(f"⚠️  验证警告: {len(all_errors)} 个（已跳过导出）")
    print("=" * 60)
    return True


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    dry_run = "--dry-run" in sys.argv

    print("=" * 60)
    print("  stick-world Excel → .tres 导出工具")
    print("=" * 60)
    print(f"  Excel 目录: {EXCEL_DIR}")
    print(f"  输出目录:   {CONFIG_DIR}")
    print(f"  模式:       {'干跑（只校验）' if dry_run else '正式导出'}")
    print("=" * 60)

    success = export_all(dry_run=dry_run)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()